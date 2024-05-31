import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.styles import Font, Alignment

# Загрузка данных из CSV
df = pd.read_csv('test_table.csv', delimiter=',', header=None,
                 names=['Группа', 'ФИО', 'Оценка1', 'Оценка2', 'Оценка3', 'Оценка4', 'Оценка5', 'Оценка6', 'Оценка7'])

# Функция для расчета оценки на основе среднего балла
def calculate_grade(avg):
    if avg >= 80:
        return 5
    elif avg >= 70:
        return 4
    elif avg >= 50:
        return 3
    elif avg >= 40:
        return 2
    else:
        return 1

# Функция для конвертации строки в float
def convert_to_float(x):
    if ',' in str(x):
        return float(str(x).replace(',', '.'))
    else:
        return float(x)

# Расчет среднего балла и оценки
df[['Оценка1', 'Оценка2', 'Оценка3', 'Оценка4', 'Оценка5', 'Оценка6', 'Оценка7']] = df[['Оценка1', 'Оценка2', 'Оценка3', 'Оценка4', 'Оценка5', 'Оценка6', 'Оценка7']].fillna('').map(convert_to_float)
df['Средний балл'] = df[['Оценка1', 'Оценка2', 'Оценка3', 'Оценка4', 'Оценка5', 'Оценка6', 'Оценка7']].mean(axis=1)
df['Оценка'] = df['Средний балл'].apply(calculate_grade)

# Создание новой книги Excel
wb = Workbook()
ws = wb.active

# Группировка по группам и создание отдельных листов
for group, data in df.groupby('Группа'):
    ws = wb.create_sheet(title=group[:31])  # Ограничение длины названия листа до 31 символа
    ws['A1'] = 'ФИО'
    ws['B1'] = 'Оценка'
    ws['C1'] = 'Средний балл'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['C1'].alignment = Alignment(horizontal='center')

    row = 2
    for _, student in data.iterrows():
        ws.cell(row=row, column=1, value=student['ФИО'])
        ws.cell(row=row, column=2, value=student['Оценка'])
        ws.cell(row=row, column=3, value=student['Средний балл'])
        row += 1

    # Добавление круговой диаграммы
    pie = PieChart()
    labels = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Распределение оценок"
    ws.add_chart(pie, "L2")

    # Добавление гистограммы
    bar = BarChart()
    data = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    bar.add_data(data, titles_from_data=True)
    bar.title = "Средняя успеваемость по заданиям"
    bar.x_axis.title = "Задание"
    bar.y_axis.title = "Средний балл"
    ws.add_chart(bar, "L16")

# Сохранение файла Excel
wb.save('1.xlsx')

print("Всё готово! Результаты уже ждут в файле 1.xlsx")